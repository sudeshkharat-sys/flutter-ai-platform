import os
import uuid
import sys
import openpyxl
from datetime import datetime

# Add backend to path to import settings if needed, 
# but we'll try to read from .env or use defaults to keep it simple
sys.path.append(os.path.join(os.getcwd(), "backend"))

try:
    from app.config import get_settings
    settings = get_settings()
    POSTGRES_CONFIG = {
        "host": settings.POSTGRES_HOST,
        "port": settings.POSTGRES_PORT,
        "user": settings.POSTGRES_USER,
        "password": settings.POSTGRES_PASSWORD,
        "database": settings.POSTGRES_DB
    }
except ImportError:
    # Fallback to manual config if app.config is not available
    POSTGRES_CONFIG = {
        "host": "localhost",
        "port": "5432",
        "user": "postgres",
        "password": "password",
        "database": "flutter_studio"
    }

EXCEL_FILE = "Model Code.xlsx"
TABLE_NAME = "master_model_mappings"

def sync_excel_to_db():
    if not os.path.exists(EXCEL_FILE):
        print(f"Error: {EXCEL_FILE} not found.")
        return

    print(f"Loading {EXCEL_FILE}...")
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    except Exception as e:
        print(f"Error loading Excel file: {e}")
        return

    import psycopg2
    try:
        conn = psycopg2.connect(**POSTGRES_CONFIG)
        cursor = conn.cursor()
    except Exception as e:
        print(f"Error connecting to PostgreSQL: {e}")
        print(f"Config used: {POSTGRES_CONFIG}")
        return

    # Ensure table exists
    cursor.execute(f"""
        CREATE TABLE IF NOT EXISTS {TABLE_NAME} (
            id VARCHAR PRIMARY KEY,
            platform_name VARCHAR NOT NULL,
            model_code VARCHAR NOT NULL UNIQUE,
            description VARCHAR,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.commit()

    total_added = 0
    total_updated = 0

    for sheet_name in wb.sheetnames:
        print(f"Processing platform: {sheet_name}")
        sheet = wb[sheet_name]
        
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
            
        header = [str(h).strip().lower() if h else "" for h in rows[0]]
        
        # Map columns
        model_code_idx = -1
        desc_idx = -1
        for i, h in enumerate(header):
            if 'model' in h and 'code' in h:
                model_code_idx = i
            if 'desc' in h:
                desc_idx = i
        
        if model_code_idx == -1: model_code_idx = 0
        if desc_idx == -1 and len(header) > 1: desc_idx = 1

        print(f"  Mapping: Model Code -> Col {model_code_idx+1}, Description -> Col {desc_idx+1}")

        for row in rows[1:]:
            if not row or len(row) <= model_code_idx:
                continue
                
            model_code = str(row[model_code_idx]).strip() if row[model_code_idx] is not None else None
            if not model_code or model_code.lower() == 'none':
                continue
                
            description = str(row[desc_idx]).strip() if desc_idx != -1 and len(row) > desc_idx and row[desc_idx] is not None else ""
            
            # Check if exists
            cursor.execute(f"SELECT id FROM {TABLE_NAME} WHERE model_code = %s", (model_code,))
            existing = cursor.fetchone()
            
            if existing:
                # Update
                cursor.execute(f"""
                    UPDATE {TABLE_NAME} 
                    SET platform_name = %s, description = %s
                    WHERE model_code = %s
                """, (sheet_name, description, model_code))
                total_updated += 1
            else:
                # Insert
                new_id = str(uuid.uuid4())
                cursor.execute(f"""
                    INSERT INTO {TABLE_NAME} (id, platform_name, model_code, description, created_at)
                    VALUES (%s, %s, %s, %s, %s)
                """, (new_id, sheet_name, model_code, description, datetime.utcnow()))
                total_added += 1

    conn.commit()
    cursor.close()
    conn.close()
    print(f"Sync complete. Added: {total_added}, Updated: {total_updated}")

if __name__ == "__main__":
    missing_deps = []
    try:
        import openpyxl
    except ImportError:
        missing_deps.append("openpyxl")
    
    try:
        import psycopg2
    except ImportError:
        missing_deps.append("psycopg2-binary")
        
    if missing_deps:
        print(f"Error: Missing dependencies: {', '.join(missing_deps)}")
        print(f"Please install them using: pip install {' '.join(missing_deps)}")
    else:
        sync_excel_to_db()
