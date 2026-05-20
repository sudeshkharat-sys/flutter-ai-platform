"""
Sync ENGINE MODELS CD LINE.xlsx → engine_model_mappings table in PostgreSQL.

Usage:
    python sync_engine_codes.py

Reads every sheet in the Excel file. For each row it tries to find the
PART NO column (looks for keywords "part", "model", "no", "code") and the
MODEL NAME / DESCRIPTION column. Inserts new rows and updates existing ones
(matched on part_no unique key).
"""

import uuid
import sys
import os
import openpyxl
import psycopg2
from psycopg2.extras import RealDictCursor

# ── DB config ─────────────────────────────────────────────────────────────────
try:
    sys.path.insert(0, os.path.join(os.path.dirname(__file__), "backend"))
    from app.config import get_settings
    s = get_settings()
    DB_CONFIG = {
        "host": s.POSTGRES_HOST,
        "port": s.POSTGRES_PORT,
        "dbname": s.POSTGRES_DB,
        "user": s.POSTGRES_USER,
        "password": s.POSTGRES_PASSWORD,
    }
except Exception:
    DB_CONFIG = {
        "host": "localhost",
        "port": 5432,
        "dbname": "flutter_ai_studio",
        "user": "postgres",
        "password": "postgres",
    }

EXCEL_FILE = os.path.join(os.path.dirname(__file__), "ENGINE MODELS CD LINE.xlsx")

# Sheets to skip (empty or not useful)
SKIP_SHEETS = {"Sheet1"}

# Column header keywords to identify columns (case-insensitive)
PART_NO_KEYWORDS   = ["part", "model", "models", "code"]
MODEL_NAME_KEYWORDS = ["model name", "model", "name"]
DESC_KEYWORDS      = ["desc", "discription", "description"]


def _find_col(headers, keywords):
    """Return index of first header that contains any keyword, else None."""
    for kw in keywords:
        for i, h in enumerate(headers):
            if h and kw.lower() in str(h).lower():
                return i
    return None


def _looks_like_code(val):
    """Return True if the value looks like an engine part number."""
    if not val:
        return False
    s = str(val).strip()
    return len(s) >= 4 and any(c.isdigit() for c in s)


def sync():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor(cursor_factory=RealDictCursor)

    added = 0
    updated = 0

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        # Find the header row (first row that has at least 2 non-null cells)
        header_row_idx = None
        for i, row in enumerate(rows):
            non_null = [c for c in row if c is not None]
            if len(non_null) >= 2:
                header_row_idx = i
                break

        if header_row_idx is None:
            print(f"  [{sheet_name}] No header row found, skipping.")
            continue

        headers = rows[header_row_idx]
        part_col  = _find_col(headers, PART_NO_KEYWORDS)
        name_col  = _find_col(headers, MODEL_NAME_KEYWORDS)
        desc_col  = _find_col(headers, DESC_KEYWORDS)

        # Fallback: part_col = first col with digit-containing values
        if part_col is None:
            part_col = 1  # col index 1 is most common in this file

        print(f"  [{sheet_name}] part_col={part_col} name_col={name_col} desc_col={desc_col}")

        data_rows = rows[header_row_idx + 1:]

        for row in data_rows:
            if not row or all(c is None for c in row):
                continue

            part_no = str(row[part_col]).strip() if part_col is not None and part_col < len(row) and row[part_col] else None
            if not _looks_like_code(part_no):
                continue

            model_name = str(row[name_col]).strip() if name_col is not None and name_col < len(row) and row[name_col] else None
            description = str(row[desc_col]).strip() if desc_col is not None and desc_col < len(row) and row[desc_col] else None

            # Check if exists
            cur.execute("SELECT id FROM engine_model_mappings WHERE part_no = %s", (part_no,))
            existing = cur.fetchone()

            if existing:
                cur.execute(
                    "UPDATE engine_model_mappings SET sheet_name=%s, model_name=%s, description=%s WHERE part_no=%s",
                    (sheet_name.upper(), model_name, description, part_no),
                )
                updated += 1
            else:
                cur.execute(
                    "INSERT INTO engine_model_mappings (id, sheet_name, part_no, model_name, description, created_at) VALUES (%s,%s,%s,%s,%s,NOW())",
                    (str(uuid.uuid4()), sheet_name.upper(), part_no, model_name, description),
                )
                added += 1

    conn.commit()
    cur.close()
    conn.close()
    print(f"\nDone — {added} added, {updated} updated.")


if __name__ == "__main__":
    print(f"Reading: {EXCEL_FILE}\n")
    sync()
